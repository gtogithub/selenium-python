import openpyxl


class TestData:
    
    data1 = [
        {"firstLetters": "pol", "countryLink": "Poland"}
        #{"firstLetters": "ita", "countryLink": "Italy"},
        #{"firstLetters": "ger", "countryLink": "Germany"}
        ]
    
    data2 = [
        {"name": "Andrew", "email": "andy@gmail.com", "gender": "Male"},
        {"name": "Bill", "email": "billy@gmail.com", "gender": "Male"},
        {"name": "Jane", "email": "jane@gmail.com", "gender": "Female"},
        {"name": "Kate", "email": "kate@gmail.com", "gender": "Female"},
        {"name": "Scooby", "email": "scooby@gmail.com", "gender": "wooof"},
        ]
    
    staticmethod
    def get_excel_data():
        file = openpyxl.load_workbook(r"C:\Users\gtoru\Desktop\Visual Studio Code\selenium python\pytest\data_for_tests.xlsx")
        sheet = file.active
        dataDict = {}
        dataList = []

        for i in range(2, sheet.max_row+1):
            for j in range(2, sheet.max_column+1):
                dataDict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            dataList.append(dataDict)
            dataDict = {}
        return dataList